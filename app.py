import json
import math
import os
import random
from datetime import date
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


APP_DIR = Path(os.getenv('APPDATA', Path.home())) / 'ClientOrderManager'
DATA_FILE = APP_DIR / 'data.json'


class Data:
    def __init__(self):
        self.value = {'clients': [], 'orders': [], 'payments': []}
        try:
            loaded = json.loads(DATA_FILE.read_text(encoding='utf-8'))
            if all(isinstance(loaded.get(x), list) for x in self.value):
                self.value = loaded
        except (OSError, ValueError):
            pass

    def save(self):
        APP_DIR.mkdir(parents=True, exist_ok=True)
        DATA_FILE.write_text(json.dumps(self.value, ensure_ascii=False, indent=2), encoding='utf-8')

    def find(self, group, ident):
        return next((x for x in self.value[group] if x['id'] == ident), None)


class Manager(tk.Tk):
    BG, PAPER, DARK, GREEN, MINT, MUTED = '#F3F6F4', '#FFFFFF', '#143C34', '#08785E', '#DFF4EB', '#63756F'

    def __init__(self):
        super().__init__()
        self.data = Data()
        self.current_page = '概览'
        self.pulse = 0
        self.quote_index = 0
        self.quotes = [
            '“我思故我在。”—— 笛卡尔', '“知识就是力量。”—— 培根', '“学而不思则罔，思而不学则殆。”—— 孔子',
            '“不积跬步，无以至千里。”—— 荀子', '“千里之行，始于足下。”—— 老子', '“天行健，君子以自强不息。”——《周易》',
            '“路漫漫其修远兮，吾将上下而求索。”—— 屈原', '“业精于勤，荒于嬉。”—— 韩愈', '“宝剑锋从磨砺出，梅花香自苦寒来。”—— 佚名',
            '“读万卷书，行万里路。”—— 刘彝', '“三人行，必有我师焉。”—— 孔子', '“己所不欲，勿施于人。”—— 孔子',
            '“知之者不如好之者，好之者不如乐之者。”—— 孔子', '“静以修身，俭以养德。”—— 诸葛亮', '“非淡泊无以明志，非宁静无以致远。”—— 诸葛亮',
            '“海纳百川，有容乃大。”—— 林则徐', '“先天下之忧而忧，后天下之乐而乐。”—— 范仲淹', '“穷则独善其身，达则兼善天下。”—— 孟子',
            '“生于忧患，死于安乐。”—— 孟子', '“人无远虑，必有近忧。”—— 孔子', '“言必信，行必果。”—— 孔子',
            '“敏而好学，不耻下问。”—— 孔子', '“吾生也有涯，而知也无涯。”—— 庄子', '“合抱之木，生于毫末。”—— 老子',
            '“为者常成，行者常至。”——《晏子春秋》', '“志不强者智不达。”—— 墨子', '“胜人者有力，自胜者强。”—— 老子',
            '“不登高山，不知天之高也。”—— 荀子', '“天下难事，必作于易。”—— 老子', '“临渊羡鱼，不如退而结网。”——《淮南子》',
            '“尽信书，则不如无书。”—— 孟子', '“不以规矩，不能成方圆。”—— 孟子', '“苟日新，日日新，又日新。”——《礼记》',
            '“人生自古谁无死，留取丹心照汗青。”—— 文天祥', '“会当凌绝顶，一览众山小。”—— 杜甫', '“长风破浪会有时，直挂云帆济沧海。”—— 李白',
            '“沉舟侧畔千帆过，病树前头万木春。”—— 刘禹锡', '“纸上得来终觉浅，绝知此事要躬行。”—— 陆游', '“横看成岭侧成峰，远近高低各不同。”—— 苏轼',
            '“不畏浮云遮望眼，只缘身在最高层。”—— 王安石', '“问渠那得清如许，为有源头活水来。”—— 朱熹', '“莫等闲，白了少年头，空悲切。”—— 岳飞',
            '“人生在勤，不索何获。”—— 张衡', '“天下兴亡，匹夫有责。”—— 顾炎武', '“落红不是无情物，化作春泥更护花。”—— 龚自珍',
            '“有志者，事竟成。”——《后汉书》', '“百闻不如一见。”——《汉书》', '“兼听则明，偏信则暗。”—— 魏征',
            '“尺有所短，寸有所长。”—— 屈原', '“工欲善其事，必先利其器。”—— 孔子', '“时间就是生命。”—— 鲁迅',
            '“其实地上本没有路，走的人多了，也便成了路。”—— 鲁迅', '“横眉冷对千夫指，俯首甘为孺子牛。”—— 鲁迅', '“世上无难事，只要肯登攀。”—— 毛泽东',
            '“为中华之崛起而读书。”—— 周恩来', '“星星之火，可以燎原。”—— 毛泽东', '“人是要有一点精神的。”—— 毛泽东',
            '“苟利国家生死以，岂因祸福避趋之。”—— 林则徐', '“青年者，人生之王，人生之春，人生之华也。”—— 李大钊', '“横空出世，莽昆仑。”—— 毛泽东',
            '“The only way to do great work is to love what you do.” — Steve Jobs', '“Stay hungry, stay foolish.” — Steve Jobs', '“Innovation distinguishes between a leader and a follower.” — Steve Jobs',
            '“Whether you think you can or you think you can’t, you’re right.” — Henry Ford', '“The future depends on what you do today.” — Mahatma Gandhi', '“It always seems impossible until it’s done.” — Nelson Mandela',
            '“Success is not final; failure is not fatal.” — Winston Churchill', '“The secret of getting ahead is getting started.” — Mark Twain', '“Believe you can and you’re halfway there.” — Theodore Roosevelt',
            '“If opportunity doesn’t knock, build a door.” — Milton Berle', '“The best way to predict the future is to create it.” — Peter Drucker', '“What we think, we become.” — Buddha',
            '“The journey of a thousand miles begins with one step.” — Lao Tzu', '“Life is really simple, but we insist on making it complicated.” — Confucius', '“Do one thing every day that scares you.” — Eleanor Roosevelt',
            '“Dream big and dare to fail.” — Norman Vaughan', '“Action is the foundational key to all success.” — Pablo Picasso', '“Quality is not an act, it is a habit.” — Aristotle',
            '“Simplicity is the ultimate sophistication.” — Leonardo da Vinci', '“Well begun is half done.” — Aristotle', '“The unexamined life is not worth living.” — Socrates',
            '“The only true wisdom is in knowing you know nothing.” — Socrates', '“What you do speaks so loudly that I cannot hear what you say.” — Emerson', '“The mind is everything. What you think you become.” — Buddha',
            '“If you can dream it, you can do it.” — Walt Disney', '“Done is better than perfect.” — Sheryl Sandberg', '“The best preparation for tomorrow is doing your best today.” — H. Jackson Brown Jr.',
            '“The harder I work, the luckier I get.” — Samuel Goldwyn', '“Start where you are. Use what you have. Do what you can.” — Arthur Ashe', '“A year from now you may wish you had started today.” — Karen Lamb',
            '“The expert in anything was once a beginner.” — Helen Hayes', '“Great things are done by a series of small things brought together.” — Van Gogh', '“Success is the sum of small efforts, repeated day in and day out.” — Robert Collier',
            '“The only limit to our realization of tomorrow is our doubts of today.” — F. D. Roosevelt', '“Happiness depends upon ourselves.” — Aristotle', '“Knowledge speaks, but wisdom listens.” — Jimi Hendrix',
            '“The purpose of our lives is to be happy.” — Dalai Lama', '“In the middle of difficulty lies opportunity.” — Albert Einstein', '“Logic will get you from A to B. Imagination will take you everywhere.” — Albert Einstein',
            '“Learn as if you will live forever, live like you will die tomorrow.” — Gandhi'
        ]
        random.shuffle(self.quotes)
        self.title('客户订单管理器 1.0')
        self.geometry('1060x690')
        self.minsize(860, 560)
        self.configure(bg=self.BG)
        self.style = ttk.Style(self)
        self.style.theme_use('clam')
        self.style.configure('Treeview', background='white', fieldbackground='white', foreground='#1C2B27', rowheight=33, font=('Microsoft YaHei', 10))
        self.style.configure('Treeview.Heading', background='#EAF2EE', foreground='#3D5951', font=('Microsoft YaHei', 10, 'bold'))
        self.style.map('Treeview', background=[('selected', '#CFEADF')], foreground=[('selected', '#143C34')])
        self.status = tk.StringVar(value='准备就绪')
        self.make_header()
        self.make_nav()
        self.content = tk.Frame(self, bg=self.BG)
        self.content.pack(fill='both', expand=True, padx=24, pady=(0, 8))
        tk.Label(self, textvariable=self.status, bg='#E7F2ED', fg='#356356', anchor='w', padx=25, pady=7, font=('Microsoft YaHei', 9)).pack(fill='x', side='bottom')
        self.show_page('概览')
        self.rotate_quote()
        self.after(50, self.animate)

    def make_header(self):
        header = tk.Frame(self, bg=self.DARK, height=94)
        header.pack(fill='x'); header.pack_propagate(False)
        title = tk.Frame(header, bg=self.DARK); title.pack(side='left', padx=26, pady=17)
        self.pulse_canvas = tk.Canvas(title, width=30, height=30, bg=self.DARK, highlightthickness=0)
        self.pulse_canvas.pack(side='left', padx=(0, 7))
        self.ring = self.pulse_canvas.create_oval(4, 4, 26, 26, outline='#5EC7A4', width=1)
        self.dot = self.pulse_canvas.create_oval(10, 10, 20, 20, fill='#B3F5DB', outline='')
        text = tk.Frame(title, bg=self.DARK); text.pack(side='left')
        tk.Label(text, text='客户订单管理器 1.0', bg=self.DARK, fg='white', font=('Microsoft YaHei', 19, 'bold')).pack(anchor='w')
        self.quote_label = tk.Label(text, text='', bg=self.DARK, fg='#B9D9CE', font=('Microsoft YaHei', 9), width=70, anchor='w')
        self.quote_label.pack(anchor='w', pady=(2,0))
        actions = tk.Frame(header, bg=self.DARK); actions.pack(side='right', padx=25)
        self.button(actions, '导出 Excel', self.export_excel, kind='header').pack(side='left', padx=4)
        self.button(actions, '导出备份', self.export_backup, kind='header').pack(side='left', padx=4)
        self.button(actions, '导入备份', self.import_backup, kind='header').pack(side='left', padx=4)

    def make_nav(self):
        nav = tk.Frame(self, bg=self.BG); nav.pack(fill='x', padx=24, pady=16)
        self.nav_buttons = {}
        for name in ('概览', '客户', '订单', '收款', '搜索'):
            b = tk.Button(nav, text=name, command=lambda n=name:self.show_page(n), relief='flat', bd=0, padx=18, pady=8, font=('Microsoft YaHei', 10, 'bold'), cursor='hand2')
            b.pack(side='left', padx=(0, 8)); self.nav_buttons[name] = b

    def button(self, parent, text, command, kind='primary'):
        colors = {'primary': (self.GREEN, 'white', '#055A48'), 'danger': ('#FBE9E9', '#9B3939', '#F3D2D2'), 'header': ('#285B50', 'white', '#37776A'), 'soft': ('#E6F1EC', '#1E6050', '#D5E8DF')}
        base, fg, hover = colors[kind]
        btn = tk.Button(parent, text=text, bg=base, fg=fg, activebackground=hover, activeforeground=fg, relief='flat', bd=0, padx=12, pady=8, font=('Microsoft YaHei', 9, 'bold'), cursor='hand2')
        def enter(_): btn.configure(bg=hover, relief='raised', bd=1)
        def leave(_): btn.configure(bg=base, relief='flat', bd=0)
        def click():
            btn.configure(bg='#063E31' if kind in ('primary','header') else hover)
            self.after(130, lambda: btn.winfo_exists() and btn.configure(bg=hover))
            self.status.set(f'已点击：{text}')
            command()
        btn.configure(command=click)
        btn.bind('<Enter>', enter); btn.bind('<Leave>', leave)
        return btn

    def animate(self):
        self.pulse = (self.pulse + 1) % 100
        w = (math.sin(self.pulse / 100 * math.tau) + 1) / 2
        outer, inner = 9 + w*4, 4 + w*3
        self.pulse_canvas.coords(self.ring, 15-outer, 15-outer, 15+outer, 15+outer)
        self.pulse_canvas.coords(self.dot, 15-inner, 15-inner, 15+inner, 15+inner)
        self.pulse_canvas.itemconfigure(self.dot, fill='#B8F8DE' if w > .5 else '#6ED1AD')
        self.after(50, self.animate)

    def rotate_quote(self):
        self.quote_label.configure(text=self.quotes[self.quote_index])
        self.quote_index = (self.quote_index + 1) % len(self.quotes)
        if self.quote_index == 0:
            random.shuffle(self.quotes)
        self.after(6500, self.rotate_quote)

    def show_page(self, name):
        self.current_page = name
        for n, b in self.nav_buttons.items():
            active = n == name
            b.configure(bg=self.GREEN if active else '#E2EAE6', fg='white' if active else '#466359', activebackground=self.GREEN if active else '#D0E0D9')
        for child in self.content.winfo_children(): child.destroy()
        getattr(self, {'概览':'page_dashboard','客户':'page_clients','订单':'page_orders','收款':'page_payments','搜索':'page_search'}[name])()

    def panel(self, title, action_text=None, command=None):
        outer = tk.Frame(self.content, bg=self.PAPER, highlightbackground='#D8E4DE', highlightthickness=1)
        head = tk.Frame(outer, bg=self.PAPER); head.pack(fill='x', padx=17, pady=12)
        tk.Label(head, text=title, bg=self.PAPER, fg='#1C302A', font=('Microsoft YaHei', 11, 'bold')).pack(side='left')
        if action_text: self.button(head, action_text, command).pack(side='right')
        body = tk.Frame(outer, bg=self.PAPER); body.pack(fill='both', expand=True, padx=17, pady=(0,15))
        return outer, body

    def tree(self, parent, columns, widths):
        wrap = tk.Frame(parent, bg=self.PAPER); wrap.pack(fill='both', expand=True)
        tree = ttk.Treeview(wrap, columns=[x[0] for x in columns], show='headings', selectmode='browse')
        for (key, label), width in zip(columns, widths):
            tree.heading(key, text=label); tree.column(key, width=width, minwidth=60, anchor='w')
        scroll = ttk.Scrollbar(wrap, command=tree.yview); tree.configure(yscrollcommand=scroll.set)
        tree.pack(side='left', fill='both', expand=True); scroll.pack(side='right', fill='y')
        return tree

    def page_dashboard(self):
        banner = tk.Frame(self.content, bg=self.DARK, height=58); banner.pack(fill='x', pady=(0,15)); banner.pack_propagate(False)
        tk.Label(banner, text='●  今天也在稳稳推进你的生意', bg=self.DARK, fg='#B5F2D7', font=('Microsoft YaHei', 11, 'bold')).pack(side='left', padx=18, pady=17)
        total = sum(float(x['amount']) for x in self.data.value['orders'])
        received = sum(float(x['amount']) for x in self.data.value['payments'])
        metrics = [('客户总数', str(len(self.data.value['clients'])), '#1C302A'), ('订单总额', self.money(total), '#1C302A'), ('已收款', self.money(received), '#08785E'), ('待收款', self.money(max(0,total-received)), '#C37A13')]
        row = tk.Frame(self.content, bg=self.BG); row.pack(fill='x', pady=(0,15))
        for title, value, color in metrics:
            card = tk.Frame(row, bg=self.PAPER, highlightbackground='#D8E4DE', highlightthickness=1)
            card.pack(side='left', fill='x', expand=True, padx=5)
            tk.Label(card, text=title, bg=self.PAPER, fg=self.MUTED, font=('Microsoft YaHei', 9)).pack(anchor='w', padx=15, pady=(13,2))
            tk.Label(card, text=value, bg=self.PAPER, fg=color, font=('Microsoft YaHei', 18, 'bold')).pack(anchor='w', padx=15, pady=(0,13))
        box, body = self.panel('待收款订单', '+ 新建订单', self.order_form); box.pack(fill='both', expand=True)
        rows = [o for o in self.data.value['orders'] if float(o.get('paid',0)) < float(o['amount'])]
        if not rows:
            tk.Label(body, text='暂无待收款订单。', bg=self.PAPER, fg=self.MUTED, font=('Microsoft YaHei', 10)).pack(pady=40); return
        t = self.tree(body, [('title','订单'),('client','客户'),('due','截止日'),('amount','待收款'),('status','状态')], [300,180,140,150,120])
        for o in rows:
            c = self.data.find('clients', o['clientId'])
            t.insert('', 'end', values=(o['title'],c['name'] if c else '—',o.get('due','—') or '—',self.money(float(o['amount'])-float(o.get('paid',0))),o.get('status','')))

    def page_clients(self):
        box, body = self.panel('客户档案', '+ 新增客户', self.client_form); box.pack(fill='both', expand=True)
        t = self.tree(body, [('name','客户'),('phone','联系方式'),('source','来源'),('status','状态'),('count','订单数'),('note','备注')], [180,150,130,105,80,320])
        for c in self.data.value['clients']:
            t.insert('', 'end', iid=c['id'], values=(c['name'],c.get('phone') or '—',c.get('source') or '—',c.get('status',''),sum(o['clientId']==c['id'] for o in self.data.value['orders']),c.get('note') or ''))
        t.bind('<Double-1>', lambda _: self.client_form(t.selection()[0]) if t.selection() else None)
        bar=tk.Frame(body,bg=self.PAPER);bar.pack(fill='x',pady=(10,0))
        self.button(bar,'编辑所选客户',lambda:self.pick(t,self.client_form)).pack(side='left')
        self.button(bar,'删除所选客户',lambda:self.delete_client(t.selection()[0]) if t.selection() else self.warn_pick(),kind='danger').pack(side='left',padx=8)

    def page_orders(self):
        box, body = self.panel('全部订单', '+ 新建订单', self.order_form); box.pack(fill='both', expand=True)
        t=self.tree(body,[('title','订单'),('client','客户'),('amount','金额'),('paid','已收'),('due','截止日'),('status','状态')],[240,160,120,120,135,100])
        for o in self.data.value['orders']:
            c=self.data.find('clients',o['clientId']);t.insert('', 'end',iid=o['id'],values=(o['title'],c['name'] if c else '—',self.money(o['amount']),self.money(o.get('paid',0)),o.get('due','—') or '—',o.get('status','')))
        t.bind('<Double-1>',lambda _:self.order_form(t.selection()[0]) if t.selection() else None)
        bar=tk.Frame(body,bg=self.PAPER);bar.pack(fill='x',pady=(10,0))
        self.button(bar,'编辑所选订单',lambda:self.pick(t,self.order_form)).pack(side='left')
        self.button(bar,'删除所选订单',lambda:self.delete_order(t.selection()[0]) if t.selection() else self.warn_pick(),kind='danger').pack(side='left',padx=8)

    def page_payments(self):
        box, body = self.panel('收款记录', '+ 记录收款', self.payment_form); box.pack(fill='both', expand=True)
        t=self.tree(body,[('date','日期'),('order','订单'),('client','客户'),('amount','金额'),('note','备注')],[130,250,180,140,320])
        for p in sorted(self.data.value['payments'],key=lambda x:x['date'],reverse=True):
            o=self.data.find('orders',p['orderId']); c=self.data.find('clients',o['clientId']) if o else None
            t.insert('', 'end',iid=p['id'],values=(p['date'],o['title'] if o else '已删除订单',c['name'] if c else '—',self.money(p['amount']),p.get('note') or ''))
        self.button(body,'删除所选收款',lambda:self.delete_payment(t.selection()[0]) if t.selection() else self.warn_pick(),kind='danger').pack(anchor='w',pady=(10,0))

    def page_search(self):
        box, body = self.panel('全局搜索'); box.pack(fill='both', expand=True)
        intro = tk.Label(body, text='可搜索客户名称、电话、来源、订单项目、备注、状态、收款日期等内容。', bg=self.PAPER, fg=self.MUTED, font=('Microsoft YaHei', 9))
        intro.pack(anchor='w', pady=(0, 10))
        toolbar = tk.Frame(body, bg=self.PAPER); toolbar.pack(fill='x', pady=(0, 12))
        query = tk.StringVar()
        entry = tk.Entry(toolbar, textvariable=query, width=43, font=('Microsoft YaHei', 11), relief='solid', bd=1)
        entry.pack(side='left', ipady=6)
        result_note = tk.Label(toolbar, text='', bg=self.PAPER, fg=self.MUTED, font=('Microsoft YaHei', 9))
        result_note.pack(side='left', padx=12)
        tree = self.tree(body, [('type','类别'), ('title','结果'), ('detail','详情'), ('related','关联信息')], [100, 260, 340, 220])

        def joined(*pieces):
            return ' '.join(str(x or '') for x in pieces).lower()

        def search():
            term = query.get().strip().lower()
            for item in tree.get_children():
                tree.delete(item)
            if not term:
                result_note.configure(text='请输入关键词后搜索')
                self.status.set('等待搜索关键词')
                return
            count = 0
            for c in self.data.value['clients']:
                if term in joined(c.get('name'), c.get('phone'), c.get('source'), c.get('status'), c.get('note')):
                    tree.insert('', 'end', values=('客户', c['name'], c.get('note') or '—', f"{c.get('phone') or '—'} · {c.get('status') or '—'}")); count += 1
            for o in self.data.value['orders']:
                client = self.data.find('clients', o['clientId'])
                if term in joined(o.get('title'), o.get('note'), o.get('status'), o.get('due'), client['name'] if client else ''):
                    tree.insert('', 'end', values=('订单', o['title'], o.get('note') or '—', f"{client['name'] if client else '—'} · {self.money(o['amount'])} · {o.get('status') or '—'}")); count += 1
            for p in self.data.value['payments']:
                order = self.data.find('orders', p['orderId'])
                client = self.data.find('clients', order['clientId']) if order else None
                if term in joined(p.get('date'), p.get('note'), p.get('amount'), order['title'] if order else '', client['name'] if client else ''):
                    tree.insert('', 'end', values=('收款', self.money(p['amount']), p.get('note') or '—', f"{p.get('date') or '—'} · {order['title'] if order else '已删除订单'}")); count += 1
            result_note.configure(text=f'找到 {count} 条结果')
            self.status.set(f'搜索完成：找到 {count} 条结果')

        self.button(toolbar, '搜索', search).pack(side='left', padx=8)
        entry.bind('<Return>', lambda _: search())
        entry.focus_set()

    def pick(self, tree, callback):
        if tree.selection(): callback(tree.selection()[0])
        else: self.warn_pick()

    def warn_pick(self): messagebox.showinfo('提示','请先选择一条记录。')
    def form(self, title, fields, values):
        win=tk.Toplevel(self);win.title(title);win.configure(bg=self.PAPER);win.transient(self);win.grab_set();win.resizable(False,False)
        frame=tk.Frame(win,bg=self.PAPER);frame.pack(padx=26,pady=22); result={}
        for i,(key,label,options) in enumerate(fields):
            tk.Label(frame,text=label,bg=self.PAPER,fg='#3D554E',font=('Microsoft YaHei',9,'bold')).grid(row=i,column=0,sticky='w',pady=6)
            var=tk.StringVar(value=str(values.get(key,'')));result[key]=var
            if options: w=ttk.Combobox(frame,textvariable=var,values=options,state='readonly',width=38)
            else: w=tk.Entry(frame,textvariable=var,width=41,font=('Microsoft YaHei',10))
            w.grid(row=i,column=1,padx=(16,0),pady=6,sticky='ew')
        return win,result

    def client_form(self, ident=None):
        old=self.data.find('clients',ident) if ident else {}
        win,v=self.form('编辑客户' if ident else '新增客户',[('name','客户名称 *',None),('phone','联系方式',None),('source','来源',None),('status','状态',['待跟进','进行中','已成交']),('note','备注',None)],old)
        if not old:v['status'].set('待跟进')
        def save():
            if not v['name'].get().strip():return messagebox.showwarning('提示','请填写客户名称。',parent=win)
            x={k:s.get().strip() for k,s in v.items()}
            if ident:self.data.find('clients',ident).update(x)
            else:self.data.value['clients'].append({'id':uuid4().hex,**x})
            self.data.save();win.destroy();self.status.set('客户已保存');self.show_page('客户')
        self.button(win,'保存客户',save).pack(anchor='e',padx=26,pady=(0,22))

    def order_form(self, ident=None):
        if not self.data.value['clients']:
            messagebox.showinfo('提示','请先新增客户。');self.client_form();return
        old=self.data.find('orders',ident) if ident else {}; choices=[f"{c['name']}｜{c['id']}" for c in self.data.value['clients']]
        selected=next((x for x in choices if x.endswith(old.get('clientId',''))),choices[0]);initial=dict(old);initial['client']=selected
        win,v=self.form('编辑订单' if ident else '新建订单',[('client','客户 *',choices),('title','项目/服务 *',None),('amount','订单金额（元）*',None),('paid','已收金额（元）',None),('due','截止日期（YYYY-MM-DD）',None),('status','订单状态',['待跟进','进行中','已完成']),('note','备注',None)],initial)
        if not old:v['paid'].set('0');v['status'].set('待跟进')
        def save():
            try: amount=float(v['amount'].get());paid=float(v['paid'].get() or 0)
            except ValueError:return messagebox.showwarning('提示','金额须为数字。',parent=win)
            if not v['title'].get().strip() or amount<0 or paid<0 or paid>amount:return messagebox.showwarning('提示','请填写项目，且已收金额必须介于 0 和订单金额之间。',parent=win)
            x={k:s.get().strip() for k,s in v.items()};x['clientId']=x.pop('client').split('｜')[-1];x['amount']=amount;x['paid']=paid
            if ident:self.data.find('orders',ident).update(x)
            else:self.data.value['orders'].append({'id':uuid4().hex,**x})
            self.data.save();win.destroy();self.status.set('订单已保存');self.show_page('订单')
        self.button(win,'保存订单',save).pack(anchor='e',padx=26,pady=(0,22))

    def payment_form(self):
        available=[o for o in self.data.value['orders'] if float(o.get('paid',0)) < float(o['amount'])]
        if not available:return messagebox.showinfo('提示','暂无可收款的订单。')
        opts=[f"{o['title']}（待收 {self.money(float(o['amount'])-float(o.get('paid',0)))})｜{o['id']}" for o in available]
        win,v=self.form('记录收款',[('order','关联订单 *',opts),('amount','收款金额（元）*',None),('date','收款日期',None),('note','备注',None)],{'order':opts[0],'date':date.today().isoformat()})
        def save():
            try: amount=float(v['amount'].get())
            except ValueError:return messagebox.showwarning('提示','金额须为数字。',parent=win)
            order=self.data.find('orders',v['order'].get().split('｜')[-1])
            if amount<=0 or amount>float(order['amount'])-float(order.get('paid',0)):return messagebox.showwarning('提示','收款金额必须大于 0 且不超过待收金额。',parent=win)
            self.data.value['payments'].append({'id':uuid4().hex,'orderId':order['id'],'amount':amount,'date':v['date'].get().strip(),'note':v['note'].get().strip()});order['paid']=round(float(order.get('paid',0))+amount,2)
            self.data.save();win.destroy();self.status.set('收款记录已保存');self.show_page('收款')
        self.button(win,'保存收款',save).pack(anchor='e',padx=26,pady=(0,22))

    def delete_client(self, ident):
        if any(o['clientId']==ident for o in self.data.value['orders']):return messagebox.showwarning('无法删除','该客户已有订单，请先删除订单。')
        if messagebox.askyesno('确认删除','确定删除该客户？'):
            self.data.value['clients']=[x for x in self.data.value['clients'] if x['id']!=ident];self.data.save();self.status.set('客户已删除');self.show_page('客户')
    def delete_order(self, ident):
        if messagebox.askyesno('确认删除','删除订单及关联收款记录？'):
            self.data.value['orders']=[x for x in self.data.value['orders'] if x['id']!=ident];self.data.value['payments']=[x for x in self.data.value['payments'] if x['orderId']!=ident];self.data.save();self.status.set('订单已删除');self.show_page('订单')
    def delete_payment(self, ident):
        p=self.data.find('payments',ident)
        if messagebox.askyesno('确认删除','确定删除此收款记录？'):
            o=self.data.find('orders',p['orderId'])
            if o:o['paid']=round(max(0,float(o.get('paid',0))-float(p['amount'])),2)
            self.data.value['payments']=[x for x in self.data.value['payments'] if x['id']!=ident];self.data.save();self.status.set('收款记录已删除');self.show_page('收款')

    def export_backup(self):
        path=filedialog.asksaveasfilename(defaultextension='.json',initialfile=f'客户订单备份-{date.today().isoformat()}.json',filetypes=[('备份文件','*.json')])
        if path:Path(path).write_text(json.dumps(self.data.value,ensure_ascii=False,indent=2),encoding='utf-8');self.status.set('备份已导出')
    def import_backup(self):
        path=filedialog.askopenfilename(filetypes=[('备份文件','*.json')])
        if not path:return
        try:
            d=json.loads(Path(path).read_text(encoding='utf-8'))
            if not all(isinstance(d.get(k),list) for k in ('clients','orders','payments')):raise ValueError
            if messagebox.askyesno('确认导入','导入会覆盖当前数据，继续吗？'):
                self.data.value=d;self.data.save();self.status.set('备份已导入');self.show_page(self.current_page)
        except Exception:messagebox.showerror('导入失败','请选择有效的备份 JSON 文件。')
    def export_excel(self):
        path=filedialog.asksaveasfilename(defaultextension='.xlsx',initialfile=f'客户订单报表-{date.today().isoformat()}.xlsx',filetypes=[('Excel 工作簿','*.xlsx')])
        if not path:return
        book=Workbook();book.remove(book.active);fill=PatternFill('solid',fgColor='08785E');font=Font(name='Microsoft YaHei',bold=True,color='FFFFFF')
        def sheet(name,heads,rows,widths,currency=()):
            ws=book.create_sheet(name);ws.append(heads)
            for c in ws[1]:c.fill=fill;c.font=font;c.alignment=Alignment(horizontal='center')
            for row in rows:ws.append(row)
            ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions;ws.row_dimensions[1].height=24
            for i,w in enumerate(widths,1):ws.column_dimensions[chr(64+i)].width=w
            for index in currency:
                for c in list(ws.columns)[index-1][1:]:c.number_format='¥#,##0.00'
        sheet('客户',['客户名称','联系方式','来源','状态','备注','订单数'],[[c['name'],c.get('phone',''),c.get('source',''),c.get('status',''),c.get('note',''),sum(o['clientId']==c['id'] for o in self.data.value['orders'])] for c in self.data.value['clients']],[20,18,16,12,38,10])
        orders=[]
        for o in self.data.value['orders']:
            c=self.data.find('clients',o['clientId']);orders.append([o['title'],c['name'] if c else '已删除客户',float(o['amount']),float(o.get('paid',0)),float(o['amount'])-float(o.get('paid',0)),o.get('due',''),o.get('status',''),o.get('note','')])
        sheet('订单',['项目/服务','客户','订单金额','已收金额','待收金额','截止日期','状态','备注'],orders,[24,18,14,14,14,14,12,35],(3,4,5))
        payments=[]
        for p in self.data.value['payments']:
            o=self.data.find('orders',p['orderId']);c=self.data.find('clients',o['clientId']) if o else None;payments.append([p.get('date',''),o['title'] if o else '已删除订单',c['name'] if c else '—',float(p['amount']),p.get('note','')])
        sheet('收款',['收款日期','订单','客户','收款金额','备注'],payments,[15,24,18,15,38],(4,))
        book.save(path);self.status.set('Excel 表格已导出');messagebox.showinfo('导出完成','Excel 表格已导出。')
    @staticmethod
    def money(value):return f'¥{float(value):,.2f}'


if __name__ == '__main__':
    Manager().mainloop()
